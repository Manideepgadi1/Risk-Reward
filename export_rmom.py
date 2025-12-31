import pandas as pd
import numpy as np
from datetime import datetime

CSV_PATH = "data.csv"
OUTPUT_FILE = f"RMom_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def calculate_metrics_with_rmom():
    """Calculate metrics including RMom for all indices"""
    df = pd.read_csv(CSV_PATH)
    
    if "DATE" not in df.columns:
        raise ValueError("Expected a 'DATE' column in the CSV.")
    
    # Parse DATE and sort
    df["DATE"] = pd.to_datetime(df["DATE"], format='%d/%m/%y', errors="coerce")
    df = df.dropna(subset=["DATE"]).sort_values("DATE")
    
    # Set DATE as index
    df = df.set_index("DATE")
    
    # All remaining columns are index price series
    price_cols = df.columns
    
    results = []
    
    print(f"Processing {len(price_cols)} indices...")
    
    for col in price_cols:
        prices = df[col].astype(float)
        
        # Drop missing prices
        non_na = prices.dropna()
        if len(non_na) < 2:
            continue
        
        p_start = non_na.iloc[0]
        p_end = non_na.iloc[-1]
        n_days = (non_na.index[-1] - non_na.index[0]).days
        if n_days <= 0 or p_start == 0:
            continue
        
        # Trading years approximation
        n_years = n_days / 365.0
        if n_years == 0:
            continue
        
        # 1. Return (CAGR)
        cagr = (p_end / p_start) ** (1.0 / n_years) - 1.0
        if not np.isfinite(cagr):
            continue
        
        # 2. Volatility (Annualized Standard Deviation)
        daily_returns = non_na.pct_change().dropna()
        if len(daily_returns) < 2:
            continue
        daily_vol = daily_returns.std(ddof=1)
        if not np.isfinite(daily_vol):
            continue
        annual_vol = daily_vol * np.sqrt(252)
        
        # 3. Risk
        risk = (annual_vol * 100) * 3.45 * 0.45
        if not np.isfinite(risk):
            continue
        
        # 4. Calculate 12-month return for momentum
        if len(non_na) >= 252:  # At least 1 year of trading data
            p_current = non_na.iloc[-1]
            p_12m_ago = non_na.iloc[-252]
            if p_12m_ago != 0:
                momentum_12m = ((p_current - p_12m_ago) / p_12m_ago) * 100
            else:
                momentum_12m = None
        else:
            momentum_12m = None
        
        if momentum_12m is not None and not np.isfinite(momentum_12m):
            momentum_12m = None
        
        results.append({
            "Index Name": col,
            "Return (%)": round(cagr * 100, 1),
            "Risk": round(risk, 1),
            "12M Momentum (%)": round(momentum_12m, 1) if momentum_12m is not None else None,
            "Momentum_12m_raw": momentum_12m  # Store raw value for RMom calculation
        })
    
    # Calculate Relative Momentum (RMom) as percentile rank
    valid_momentum_data = [(i, r['Momentum_12m_raw']) for i, r in enumerate(results) if r['Momentum_12m_raw'] is not None]
    
    if len(valid_momentum_data) > 1:
        # Sort by momentum value to get rankings
        sorted_momentum = sorted(valid_momentum_data, key=lambda x: x[1])
        
        # Assign percentile rank (0-100 scale)
        for rank, (original_idx, momentum_value) in enumerate(sorted_momentum):
            n = len(sorted_momentum)
            if n > 1:
                percentile = (rank / (n - 1)) * 100
            else:
                percentile = 50
            
            results[original_idx]['RMom'] = round(percentile, 1)
    else:
        for result in results:
            result['RMom'] = None
    
    # Remove temporary field
    for result in results:
        del result['Momentum_12m_raw']
    
    return results

# Calculate metrics
print("Calculating metrics and RMom...")
data = calculate_metrics_with_rmom()

# Convert to DataFrame
df_export = pd.DataFrame(data)

# Reorder columns
df_export = df_export[['Index Name', 'Return (%)', 'Risk', '12M Momentum (%)', 'RMom']]

# Sort by RMom (descending - higher is better)
df_export = df_export.sort_values('RMom', ascending=False, na_position='last')

# Export to Excel with formatting
print(f"Exporting to {OUTPUT_FILE}...")
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    df_export.to_excel(writer, sheet_name='RMom Data', index=False)
    
    # Get the worksheet
    worksheet = writer.sheets['RMom Data']
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 40  # Index Name
    worksheet.column_dimensions['B'].width = 12  # Return
    worksheet.column_dimensions['C'].width = 12  # Risk
    worksheet.column_dimensions['D'].width = 18  # 12M Momentum
    worksheet.column_dimensions['E'].width = 12  # RMom
    
    # Format header row
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        # Center align numeric columns
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

print(f"✓ Successfully exported {len(df_export)} indices to {OUTPUT_FILE}")
print(f"\nTop 10 indices by RMom:")
print(df_export.head(10).to_string(index=False))
print(f"\nBottom 10 indices by RMom:")
print(df_export.tail(10).to_string(index=False))
